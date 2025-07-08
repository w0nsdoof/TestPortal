"""
kelet_importer.py
-----------------
Robust parser + analyser + importer for KELET test workbooks.

Assumptions
-----------
* KELET workbooks are XLSX (Excel 2007+) and named `KELET-<Level>.xlsx`
  where <Level> ∈ {A1, A2, B1, B2, C1}.
* Correct answers are indicated with **bold** font in the option cell.
* Each question block begins with a cell whose text starts with "n."
  (e.g. "1."  "25.") in column B; the actual prompt is normally in column C.
* Four options A–D follow immediately below in column B/C.
* Reading‑type questions may have an optional *paragraph* cell above the prompt.
  If so, it is stored.

If any sheet deviates from the expected pattern the code **logs** the issue
and continues, instead of aborting.
"""

import os
import django

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")
django.setup()

import json
import logging
import re
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Optional

from openpyxl import load_workbook
from django.db import transaction, IntegrityError
from django.utils.text import slugify

# ───────────────────────────────────────────────────────────────────────────────
# Configure logging
# ───────────────────────────────────────────────────────────────────────────────
logging.basicConfig(
    format="%(levelname)s | %(message)s",
    level=logging.INFO,
)
log = logging.getLogger(__name__)

# ───────────────────────────────────────────────────────────────────────────────
# Data classes (for temporary in‑memory representation)
# ───────────────────────────────────────────────────────────────────────────────
@dataclass
class OptionDTO:
    label: str
    text: str
    is_correct: bool


@dataclass
class QuestionDTO:
    level: str
    qtype: str  # Grammar | Vocabulary | Reading
    prompt: str
    options: List[OptionDTO]
    paragraph: Optional[str] = None
    source: Optional[str] = None  # sheet & row for debugging


# ───────────────────────────────────────────────────────────────────────────────
# Parser
# ───────────────────────────────────────────────────────────────────────────────
QUESTION_NUMBER = re.compile(r"^\d+\.")

PREP_LIST = re.compile(
    r"([A-Z]{2,})(?:\s*[–/,]\s*([A-Z]{2,}))+",
    flags=re.I,
)

def _extract_preposition_options(text: str) -> Optional[List[str]]:
    """Return list of prepositions if the prompt contains a list like
    'AT – FOR - IN – OF – ON – TO - WITH'."""
    if not text:
        return None
    # dash variants: – — - /
    parts = re.split(r"\s*[–—,/]\s*", text)
    parts = [p.strip() for p in parts if p.strip().isalpha() and len(p) <= 5]
    return parts if 3 <= len(parts) <= 20 else None


def parse_sheet(ws, level: str, qtype: str) -> List[QuestionDTO]:
    questions: List[QuestionDTO] = []
    r = 1
    while r <= ws.max_row:
        num_cell = ws.cell(row=r, column=2)  # “1.” etc.
        if num_cell.value and isinstance(num_cell.value, str) \
           and QUESTION_NUMBER.match(num_cell.value.strip()):

            prompt_cell = ws.cell(row=r, column=3)
            prompt_txt = str(prompt_cell.value or "").strip()

            # ── Branch A: “preposition list” format ──────────────────────────
            preps = _extract_preposition_options(prompt_txt)
            if preps and qtype.lower() == "grammar":
                # Look ahead ≤5 rows in column C for single‑word answer
                answer = None
                for look in range(1, 6):
                    ans_txt = str(ws.cell(row=r + look, column=3).value or "").strip()
                    if ans_txt.isalpha() and 1 <= len(ans_txt) <= 5:
                        answer = ans_txt.upper()
                        break
                if not answer:
                    log.warning("No explicit answer for preposition Q at R%s", r)

                opts = []
                for idx, word in enumerate(preps):
                    opts.append(
                        OptionDTO(
                            label=chr(65 + idx),         # A/B/C…
                            text=word,
                            is_correct=(word.upper() == answer),
                        )
                    )

                questions.append(
                    QuestionDTO(
                        level=level,
                        qtype=qtype,
                        prompt=prompt_txt,
                        options=opts,
                        paragraph=None,
                        source=f"{ws.title} R{r}",
                    )
                )
                r += look + 1  # skip past answer row
                continue
            # ── Branch B: “normal A–D” format (unchanged) ───────────────────
            paragraph = None
            if (
                r > 2
                and ws.cell(row=r - 1, column=3).value
                and not QUESTION_NUMBER.match(str(ws.cell(row=r - 1, column=2).value or ""))
            ):
                paragraph = str(ws.cell(row=r - 1, column=3).value).strip()

            options: List[OptionDTO] = []
            for i in range(4):
                label_cell = ws.cell(row=r + 1 + i, column=2)
                text_cell = ws.cell(row=r + 1 + i, column=3)
                if not (label_cell.value and text_cell.value):
                    continue
                options.append(
                    OptionDTO(
                        label=str(label_cell.value).strip().replace(".", ""),
                        text=str(text_cell.value).strip(),
                        is_correct=bool(text_cell.font and text_cell.font.bold),
                    )
                )

            questions.append(
                QuestionDTO(
                    level=level,
                    qtype=qtype,
                    prompt=prompt_txt,
                    options=options,
                    paragraph=paragraph,
                    source=f"{ws.title} R{r}",
                )
            )
            r += 6
        else:
            r += 1
    return questions


def parse_workbook(xlsx_path: Path) -> List[QuestionDTO]:
    level = xlsx_path.stem.split("-")[-1]  # e.g. KELET-B1 → B1
    wb = load_workbook(xlsx_path, data_only=True)
    all_qs: List[QuestionDTO] = []
    mapping = {
        "grammar": "Grammar",
        "vocabulary": "Vocabulary",
        "reading": "Reading",
    }
    for ws in wb.worksheets:
        lower = ws.title.lower()
        for key, qtype in mapping.items():
            if key in lower:
                all_qs.extend(parse_sheet(ws, level, qtype))
                break
    log.info("Parsed %s → %d questions", xlsx_path.name, len(all_qs))
    return all_qs


# ───────────────────────────────────────────────────────────────────────────────
# Analytics helper
# ───────────────────────────────────────────────────────────────────────────────
def analyse(qs: List[QuestionDTO]) -> dict:
    """Return a dict of useful stats (duplicates, field lengths, etc.)."""
    duplicates = Counter(q.prompt for q in qs)
    dupes = {k: v for k, v in duplicates.items() if v > 1}

    max_prompt_len = max(len(q.prompt) for q in qs)
    max_opt_len = max(len(opt.text) for q in qs for opt in q.options)

    option_distribution = Counter(len(q.options) for q in qs)
    missing_answer_qs = [q for q in qs if not any(o.is_correct for o in q.options)]

    return {
        "total": len(qs),
        "duplicates": dupes,
        "max_prompt_len": max_prompt_len,
        "max_option_len": max_opt_len,
        "option_distribution": option_distribution,
        "missing_correct": len(missing_answer_qs),
    }


# ───────────────────────────────────────────────────────────────────────────────
# Django‑side: suggested model code generator  (pure strings)
# ───────────────────────────────────────────────────────────────────────────────
MODEL_TEMPLATE = """
class Question(models.Model):
    TYPE_CHOICES = [
        ("Grammar", "Grammar"),
        ("Vocabulary", "Vocabulary"),
        ("Reading", "Reading"),
    ]
    LEVEL_CHOICES = [
        ("A1", "A1"), ("A2", "A2"),
        ("B1", "B1"), ("B2", "B2"),
        ("C1", "C1"),
    ]
    type     = models.CharField(max_length=9, choices=TYPE_CHOICES)
    level    = models.CharField(max_length=2, choices=LEVEL_CHOICES)
    prompt   = models.CharField(max_length={prompt_len})
    paragraph = models.TextField(blank=True, null=True)
    created_at = models.DateTimeField(auto_now_add=True)

class Option(models.Model):
    question   = models.ForeignKey(Question, related_name="options", on_delete=models.CASCADE)
    label      = models.CharField(max_length=1)
    text       = models.CharField(max_length={option_len})
    is_correct = models.BooleanField(default=False)
""".lstrip()


def suggest_models(stats: dict) -> str:
    """Return a string with `models.py` fragment using calculated field sizes."""
    return MODEL_TEMPLATE.format(
        prompt_len=max(255, stats["max_prompt_len"] + 10),  # guard against >255
        option_len=max(255, stats["max_option_len"] + 10),
    )


# ───────────────────────────────────────────────────────────────────────────────
# Import into existing models (expects you already have Question & Option)
# Adjust the import paths below to match your project.
# ───────────────────────────────────────────────────────────────────────────────
def import_into_django(qs: List[QuestionDTO]):
    from questions.models import Question, Option  # <-- adjust if different
    from django.db.models import Q

    stats = analyse(qs)
    log.info("Importing %d questions…", stats["total"])

    with transaction.atomic():
        for q in qs:
            try:
                # De‑duplicate by prompt + level + type
                obj, created = Question.objects.get_or_create(
                    level=q.level,
                    type=q.qtype,
                    prompt=q.prompt,
                    defaults={"paragraph": q.paragraph},
                )
                if created:
                    for opt in q.options:
                        Option.objects.create(
                            question=obj,
                            label=opt.label,
                            text=opt.text,
                            is_correct=opt.is_correct,
                        )
                else:
                    log.debug("Skipped duplicate: %s", q.prompt[:40])
            except IntegrityError as exc:
                log.error("DB error for '%s' (%s): %s", q.prompt[:40], q.source, exc)
                continue

    log.info("Done. Missing‑answer questions skipped: %s", stats["missing_correct"])


# ───────────────────────────────────────────────────────────────────────────────
# “Main” section: parse → analyse → generate models snippet → import
# ───────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    DATA_DIR = Path(r"C:\Coding\KBTU\KELET_TESTPORTAL\backend\data")  # <- change to your folder
    all_questions: List[QuestionDTO] = []
    for xlsx in sorted(DATA_DIR.glob("KELET-*.xlsx")):
        try:
            all_questions.extend(parse_workbook(xlsx))
        except Exception as e:  # catch any parsing failure, keep going
            log.exception("Failed to parse %s: %s", xlsx.name, e)

    if not all_questions:
        log.error("No questions parsed – aborting.")
        exit(1)

    # ── 1. analytics / trends
    stats = analyse(all_questions)
    log.info("GLOBAL STATS: %s", json.dumps(stats, indent=2, default=str))

    # ── 2. suggest model fragment (print once for convenience)
    print("\n# ── Suggested models.py fragment ─────────────────────────────")
    print(suggest_models(stats))
    print("# ──────────────────────────────────────────────────────────────\n")

    # ── 3. import into DB (comment out if you only want analysis)
    import_into_django(all_questions)

    log.info("Script finished.")
