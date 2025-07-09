import re
import os
import openpyxl
from .models import Question, Option, QuestionType

# === Dispatcher ===
def import_questions_from_excel(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    level = os.path.splitext(os.path.basename(file_path))[0][-2:]

    sheet_parser_map = {
        "grammar": parse_grammar_sheet,
        "vocabulary": parse_vocabulary_sheet,
        "reading": parse_reading_sheet,
    }

    for sheet_name in wb.sheetnames:
        matched = next((key for key in sheet_parser_map if key in sheet_name.lower()), None)

        if matched:
            print(f"📄 Parsing sheet: {sheet_name} as {matched}")
            sheet_parser_map[matched](wb[sheet_name], level)
        else:
            print(f"⚠️ Skipping unknown sheet: {sheet_name}")


def is_numbered(cell_value):
    return isinstance(cell_value, str) and re.match(r"^\d+\.", cell_value.strip())

def parse_grammar_sheet(sheet, level):
    row = 1
    while row < sheet.max_row:
        number_cell = sheet.cell(row=row, column=2).value
        text_cell = sheet.cell(row=row, column=3).value

        if is_numbered(number_cell) and isinstance(text_cell, str):
            instruction = text_cell.strip()
            question_row = row + 1
            question_text = sheet.cell(row=question_row, column=3).value

            if not question_text:
                row += 1
                continue

            # ===== Case 1: Fill-in-the-gap with prepositions (options in instruction)
            if instruction.lower().startswith("complete the sentence"):
                answer_row = question_row + 2
                answer = (sheet.cell(row=answer_row, column=3).value or "").strip().lower()
                option_list = re.split(r"[–\-•,]", instruction.split(":")[-1])
                option_list = [opt.strip().lower() for opt in option_list if opt.strip()]

                q = Question.objects.create(
                    type=QuestionType.GRAMMAR,
                    level=level,
                    prompt=question_text
                )

                # Check for bold options
                use_bold = False
                bold_indices = set()
                for idx, opt in enumerate(option_list):
                    opt_cell = sheet.cell(row=question_row + 2, column=3)  # No per-option cell for this type
                    # For this type, we can't check bold per option, so fallback to answer logic
                for idx, opt in enumerate(option_list):
                    Option.objects.create(
                        question=q,
                        label=chr(65 + idx),
                        text=opt,
                        is_correct=(opt == answer)
                    )
                row += 4

            # ===== Case 2: MCQ with labeled answers below
            elif instruction.lower().startswith("choose the appropriate answer"):
                options = []
                option_cells = []
                correct_index = None
                bold_indices = set()
                for i, label in enumerate(['A', 'B', 'C', 'D']):
                    opt_cell = sheet.cell(row=row + 2 + i, column=3)
                    option_text = opt_cell.value
                    if not option_text:
                        continue
                    option_text = option_text.strip()
                    options.append((label, option_text))
                    option_cells.append(opt_cell)
                    if opt_cell.font and opt_cell.font.bold:
                        bold_indices.add(i)
                    if option_text.lower().startswith("it seems"):  # Example heuristic
                        correct_index = i
                use_bold = len(bold_indices) > 0
                q = Question.objects.create(
                    type=QuestionType.GRAMMAR,
                    level=level,
                    prompt=question_text
                )
                for i, (label, opt_text) in enumerate(options):
                    if use_bold:
                        is_correct = i in bold_indices
                    else:
                        is_correct = (i == correct_index)
                    Option.objects.create(
                        question=q,
                        label=label,
                        text=opt_text,
                        is_correct=is_correct
                    )
                row += 6
            else:
                row += 1
        else:
            row += 1

def parse_reading_sheet(sheet, level):
    row = 1
    while row <= sheet.max_row:
        number_cell = sheet.cell(row=row, column=2).value
        if number_cell and re.match(r"^\d+\.", str(number_cell).strip()):
            instruction = sheet.cell(row=row, column=3).value
            instruction = str(instruction).strip() if instruction else ""
            if re.search(r"read the (paragraph|article)", instruction, re.IGNORECASE):
                row += 1
                paragraph_lines = []
                while row <= sheet.max_row:
                    para_text = sheet.cell(row=row, column=3).value
                    if para_text is None:
                        row += 1
                        continue
                    para_text_str = str(para_text).strip()
                    if re.match(r"^(Complete the sentence:|Question:|Answer the questions\.)", para_text_str, re.IGNORECASE):
                        break
                    if para_text_str:
                        paragraph_lines.append(para_text_str)
                    row += 1
                paragraph = "\n".join(paragraph_lines)
                question_text = ""
                if row <= sheet.max_row:
                    block_marker = sheet.cell(row=row, column=3).value
                    block_marker_str = str(block_marker).strip() if block_marker else ""
                    m = re.match(r"^(Complete the sentence:|Question:|Answer the question\.)(.*)", block_marker_str, re.IGNORECASE)
                    if m and m.group(2).strip():
                        question_text = m.group(2).strip()
                        row += 1
                    else:
                        row += 1
                        while row <= sheet.max_row:
                            next_text = sheet.cell(row=row, column=3).value
                            if next_text and str(next_text).strip():
                                question_text = str(next_text).strip()
                                row += 1
                                break
                            row += 1
                q = Question.objects.create(
                    type=QuestionType.READING,
                    level=level,
                    prompt=f"{instruction}\n\n{question_text}",
                    paragraph=paragraph
                )
                option_labels = ['a', 'b', 'c', 'd', 'e']
                option_cells = []
                bold_indices = set()
                options = []
                start_row = row
                for idx, label in enumerate(option_labels):
                    if row > sheet.max_row:
                        break
                    opt_label = sheet.cell(row=row, column=2).value
                    opt_cell = sheet.cell(row=row, column=3)
                    opt_text = opt_cell.value
                    if (opt_label is None or str(opt_label).strip().lower() != label) or not opt_text:
                        break
                    opt_text = str(opt_text).strip()
                    options.append((label, opt_text))
                    option_cells.append(opt_cell)
                    if opt_cell.font and opt_cell.font.bold:
                        bold_indices.add(idx)
                    row += 1
                use_bold = len(bold_indices) > 0
                for idx, (label, opt_text) in enumerate(options):
                    is_correct = idx in bold_indices if use_bold else False
                    Option.objects.create(
                        question=q,
                        label=label.upper(),
                        text=opt_text,
                        is_correct=is_correct
                    )
            else:
                row += 1
        else:
            row += 1

def parse_vocabulary_sheet(sheet, level):
    row = 1
    while row < sheet.max_row:
        number_cell = sheet.cell(row=row, column=2).value
        text_cell = sheet.cell(row=row, column=3).value

        if is_numbered(number_cell) and isinstance(text_cell, str):
            instruction = text_cell.strip()
            # === Type 1: Complete... (options in instruction)
            if ":" in instruction:
                options_part = instruction.split(":")[-1]
                option_list = [opt.strip().lower() for opt in re.split(r"[\\/]", options_part) if opt.strip()]
                question_row = row + 1
                question_text = sheet.cell(row=question_row, column=3).value
                if not question_text:
                    row += 1
                    continue
                question_text = question_text.strip()
                answer_row = row + 3
                answer = (sheet.cell(row=answer_row, column=3).value or "").strip().lower()
                q = Question.objects.create(
                    type=QuestionType.VOCABULARY,
                    level=level,
                    prompt=question_text
                )
                for idx, opt in enumerate(option_list):
                    Option.objects.create(
                        question=q,
                        label=chr(65 + idx),  # A, B, C, ...
                        text=opt,
                        is_correct=(opt == answer)
                    )
                row += 4
            # === Type 2: Choose... (options below, correct is bold)
            elif instruction.lower().startswith("choose"):
                question_row = row + 1
                question_text = sheet.cell(row=question_row, column=3).value
                if not question_text:
                    row += 1
                    continue
                question_text = question_text.strip()
                # Read options (a, b, c, d, ...)
                options = []
                correct_indices = set()
                option_labels = ['a', 'b', 'c', 'd', 'e']
                opt_row = question_row + 1
                while opt_row <= sheet.max_row:
                    opt_label = sheet.cell(row=opt_row, column=2).value
                    opt_cell = sheet.cell(row=opt_row, column=3)
                    opt_text = opt_cell.value
                    if opt_label is None or str(opt_label).strip().lower() not in option_labels or not opt_text:
                        break
                    label = str(opt_label).strip().lower()
                    text = str(opt_text).strip()
                    is_bold = opt_cell.font and opt_cell.font.bold
                    options.append((label, text, is_bold))
                    if is_bold:
                        correct_indices.add(len(options) - 1)
                    opt_row += 1
                # If any option is bold, use bold for correctness; else, use answer row
                use_bold = any(is_bold for _, _, is_bold in options)
                answer = None
                answer_row = opt_row
                if not use_bold:
                    while answer_row <= sheet.max_row:
                        ans_val = sheet.cell(row=answer_row, column=3).value
                        if ans_val and str(ans_val).strip():
                            answer = str(ans_val).strip().lower()
                            break
                        answer_row += 1
                q = Question.objects.create(
                    type=QuestionType.VOCABULARY,
                    level=level,
                    prompt=question_text
                )
                for idx, (label, opt_text, is_bold) in enumerate(options):
                    if use_bold:
                        is_correct = is_bold
                    else:
                        is_correct = (answer == label) or (answer == opt_text.lower())
                    Option.objects.create(
                        question=q,
                        label=label.upper(),
                        text=opt_text,
                        is_correct=is_correct
                    )
                row = answer_row + 1 if not use_bold else opt_row
            else:
                row += 1
        else:
            row += 1


def import_questions_from_json(json_data, level=None, clear_existing=False):
    """
    Import questions from JSON data
    
    Args:
        json_data (list): List of question dictionaries
        level (str): English level (A1, A2, B1, B2, C1)
        clear_existing (bool): Whether to clear existing questions for this level
    
    Returns:
        int: Number of questions imported
    """
    from django.db import transaction
    
    if not isinstance(json_data, list):
        raise ValueError("JSON data must be a list of questions")
    
    if not level:
        raise ValueError("Level must be specified")
    
    with transaction.atomic():
        if clear_existing:
            Question.objects.filter(level=level).delete()
        
        imported_count = 0
        
        for question_data in json_data:
            try:
                # Extract question fields
                question_type = question_data.get('type', '').upper()
                prompt = question_data.get('prompt', '')
                paragraph = question_data.get('paragraph', '')
                options_data = question_data.get('options', [])
                
                # Validate required fields
                if not prompt or not options_data:
                    continue
                
                # Map question type
                type_mapping = {
                    'GRAMMAR': QuestionType.GRAMMAR,
                    'READING': QuestionType.READING,
                    'VOCABULARY': QuestionType.VOCABULARY,
                }
                
                if question_type not in type_mapping:
                    continue
                
                # Create the question
                question = Question.objects.create(
                    type=type_mapping[question_type],
                    level=level,
                    prompt=prompt,
                    paragraph=paragraph if paragraph else None
                )
                
                # Create options
                for option_data in options_data:
                    label = option_data.get('label', '')
                    text = option_data.get('text', '')
                    is_correct = option_data.get('is_correct', False)
                    
                    if not text:
                        continue
                    
                    Option.objects.create(
                        question=question,
                        label=label.upper(),
                        text=text,
                        is_correct=is_correct
                    )
                
                imported_count += 1
                
            except Exception as e:
                print(f"Error creating question: {e}")
                continue
        
        return imported_count


def import_questions_from_json_file(file_path, level=None, clear_existing=False):
    """
    Import questions from a JSON file
    
    Args:
        file_path (str): Path to the JSON file
        level (str): English level (A1, A2, B1, B2, C1)
        clear_existing (bool): Whether to clear existing questions for this level
    
    Returns:
        int: Number of questions imported
    """
    import json
    import os
    
    # Determine level from filename if not provided
    if not level:
        filename = os.path.basename(file_path)
        if filename.startswith('a2'):
            level = 'A2'
        elif filename.startswith('b1'):
            level = 'B1'
        elif filename.startswith('b2'):
            level = 'B2'
        elif filename.startswith('c1'):
            level = 'C1'
        else:
            level = 'A1'  # Default
    
    with open(file_path, 'r', encoding='utf-8') as file:
        json_data = json.load(file)
    
    return import_questions_from_json(json_data, level, clear_existing)


